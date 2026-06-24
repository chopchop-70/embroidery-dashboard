"""
generate_dashboard_data.py
Reads the two Excel files, builds data.json, and pushes it to GitHub
so the Vercel dashboard always has fresh data.

Run automatically by Task Scheduler alongside embroidery_final.py.
Usage: python C:\Scripts\generate_dashboard_data.py
"""
import json
import re
import requests
import base64
from datetime import datetime
from openpyxl import load_workbook

# ── Config ────────────────────────────────────────────────────────────────────
PATH_4 = r"C:\Embroidery\Embroidery - 4 head.xlsx"
PATH_6 = r"C:\Embroidery\Embroidery - 6 head.xlsx"

GITHUB_TOKEN = "ghp_i7j9JDjbx0L0Q8zz0ykQ38T2jdwcu01x2QuI"   # personal access token (repo scope)
GITHUB_REPO  = "chopchop-70/embroidery-dashboard"  # e.g. jasondacey/embroidery-dashboard
GITHUB_FILE  = "public/data.json"         # path inside the repo
GITHUB_BRANCH = "main"

# ── Date parsing ──────────────────────────────────────────────────────────────
MONTH_MAP = {'Jan':1,'Feb':2,'Mar':3,'Apr':4,'May':5,'Jun':6,
             'Jul':7,'Aug':8,'Sep':9,'Oct':10,'Nov':11,'Dec':12}
SKIP_VALS = {'DATE','REPORTS STOPPED','Friday 3.4.26 Public Holiday',
             'Monday 6.4.26 Public Holiday',''}

def parse_date(raw):
    s = str(raw).strip().lstrip(': ')
    m = re.search(r'(\w+)\s+(\w+)\s+(\d+)\s+\d+:\d+:\d+\s+(\d{4})', s)
    if m:
        mon, day, yr = m.group(2), int(m.group(3)), int(m.group(4))
        if mon in MONTH_MAP:
            return datetime(yr, MONTH_MAP[mon], day)
    m2 = re.match(r'(\d{1,2})/(\d{1,2})/(\d{4})', s)
    if m2:
        return datetime(int(m2.group(3)), int(m2.group(2)), int(m2.group(1)))
    return None

def to_mins(s):
    m = re.search(r'(\d+):(\d+):(\d+)', str(s))
    return int(m.group(1))*60+int(m.group(2)) if m else 0

# ── Summary extraction ────────────────────────────────────────────────────────
def extract_summary(path, sheet, machine):
    wb = load_workbook(path, data_only=True)
    ws = wb[sheet]
    records = []
    for row in ws.iter_rows(min_row=2):
        a = row[0].value
        if not a or str(a).strip() in SKIP_VALS:
            continue
        dt = parse_date(str(a))
        if not dt:
            continue
        try:
            eff = float(str(row[5].value).replace('%','').strip()) if row[5].value else 0
            if eff <= 0:
                continue
            records.append({
                "date":          dt.strftime('%b %d %Y'),
                "date_sort":     dt.strftime('%Y-%m-%d'),
                "machine":       machine,
                "efficiency":    round(eff, 2),
                "downtime_min":  to_mins(row[2].value),
                "runtime_min":   to_mins(row[3].value),
                "thread_breaks": int(float(str(row[8].value)))  if row[8].value  else 0,
                "pieces":        int(float(str(row[11].value))) if row[11].value else 0,
            })
        except Exception as e:
            print(f"  Skipping summary row {row[0].row}: {e}")
    return records

# ── Design extraction ─────────────────────────────────────────────────────────
def extract_designs(path, sheet, machine):
    wb = load_workbook(path, data_only=True)
    ws = wb[sheet]
    records, current_date = [], None
    for row in ws.iter_rows(min_row=2):
        a, b = row[0].value, row[1].value
        if a and str(a).strip() not in ('DATE','None',''):
            dt = parse_date(str(a))
            if dt:
                current_date = dt.strftime('%b %d %Y')
        if not b or str(b).strip() in ('Design','None',''):
            continue
        if not current_date:
            continue
        try:
            runs = int(float(str(row[2].value))) if row[2].value else 0
            if runs <= 0:
                continue
            breaks = int(float(str(row[3].value))) if row[3].value else 0
            t = str(row[4].value) if row[4].value else ''
            tm = re.search(r'(\d+):(\d+):(\d+)', t)
            avg_mins = round(int(tm.group(1))*60+int(tm.group(2))+int(tm.group(3))/60, 2) if tm else 0
            records.append({
                "date":         current_date,
                "machine":      machine,
                "design":       str(b).strip(),
                "runs":         runs,
                "thread_breaks": breaks,
                "avg_run_mins": avg_mins,
            })
        except Exception as e:
            print(f"  Skipping design row {row[0].row}: {e}")
    return records

# ── GitHub push ───────────────────────────────────────────────────────────────
def push_to_github(content_str):
    url = f"https://api.github.com/repos/{GITHUB_REPO}/contents/{GITHUB_FILE}"
    headers = {
        "Authorization": f"token {GITHUB_TOKEN}",
        "Accept": "application/vnd.github.v3+json",
    }
    # Get current file SHA (needed for update)
    sha = None
    r = requests.get(url, headers=headers, params={"ref": GITHUB_BRANCH})
    if r.status_code == 200:
        sha = r.json().get("sha")

    payload = {
        "message": f"Update dashboard data {datetime.now().strftime('%Y-%m-%d %H:%M')}",
        "content": base64.b64encode(content_str.encode()).decode(),
        "branch":  GITHUB_BRANCH,
    }
    if sha:
        payload["sha"] = sha

    r = requests.put(url, headers=headers, json=payload)
    if r.status_code in (200, 201):
        print(f"  GitHub push OK — {r.json()['commit']['sha'][:8]}")
    else:
        print(f"  GitHub push FAILED: {r.status_code} {r.text[:200]}")

# ── Main ──────────────────────────────────────────────────────────────────────
def main():
    print("Generating dashboard data...")

    s6 = extract_summary(PATH_6, "6 SUMMARY(Sheet1)", "6-head")
    s4 = extract_summary(PATH_4, "4 SUMMARY(Sheet1)", "4-head")
    d6 = extract_designs(PATH_6, "6 Design Summary", "6-head")
    d4 = extract_designs(PATH_4, "4 Design Summary", "4-head")

    print(f"  Summary: {len(s6)} 6-head + {len(s4)} 4-head rows")
    print(f"  Designs: {len(d6)} 6-head + {len(d4)} 4-head rows")

    data = {
        "generated": datetime.now().isoformat(timespec='seconds'),
        "summary":   sorted(s6 + s4, key=lambda r: r["date_sort"]),
        "designs":   d6 + d4,
    }

    content = json.dumps(data, separators=(',',':'))
    print(f"  JSON size: {len(content):,} bytes")

    if GITHUB_TOKEN == "YOUR_GITHUB_TOKEN_HERE":
        print("\nWARNING: GitHub token not configured — saving locally only.")
        with open(r"C:\Scripts\data.json", "w") as f:
            f.write(content)
        print("  Saved to C:\\Scripts\\data.json")
    else:
        print("  Pushing to GitHub...")
        push_to_github(content)

    print("Done.")

if __name__ == "__main__":
    main()
